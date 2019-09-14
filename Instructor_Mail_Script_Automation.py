import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import Alignment, Font

# input, output 파일 세팅
INPUT_PARA_FILE = 'input_parameters.xlsx'
OUTPUT_SCRIPT_FILE = 'mail_script.xlsx'

input_para_workbook = openpyxl.load_workbook(INPUT_PARA_FILE)
input_para_worksheet = input_para_workbook['스크립트']

# 전역변수 설정
EDU_SECTION = input_para_worksheet['B4'].value
GREETING_START = input_para_worksheet['B5'].value
GREETING_END = input_para_worksheet['B6'].value
SCHEDULE_FILE = input_para_worksheet['B7'].value

EDU_SECTION_COLUMN = 'A'
EDU_COLUMN = 'B'
EDU_ROOM_COLUMN = 'C'
CHANGED_EDU_ROOM_COLUMN = 'D'
TIME_COLUMN = 'J'
MONDAY_COLUMN = 'K'
FIRST_WEEK_ROW = 4

END_ROW_VALUE = '강사별 투입일수("공우식", "공우식/", "/공우식", "공우식?", "공우식?/", "/공우식?" 까지만 인식)'

DAY_OF_WEEK_COLUMNS = ['K', 'L', 'M', 'N', 'O', 'P', 'Q']
WEEKDAY_COLUMNS = ['K', 'L', 'M', 'N', 'O']

DETAILED_EDU_TIME = {
    "7": ["09:30 ~ 17:30"],
    "7-7": ["09:30 ~ 17:30", "09:30 ~ 17:30"],
    "7-7-4": ["09:30 ~ 17:30", "09:30 ~ 17:30", "09:00 ~ 13:00"],
    "4-7-7": ["14:00 ~ 18:00", "09:30 ~ 17:30", "09:30 ~ 17:30"],
    "6-6-6": ["10:00 ~ 17:00", "10:00 ~ 17:00", "10:00 ~ 17:00"],
    "7-7-7": ["09:30 ~ 17:30", "09:30 ~ 17:30", "09:30 ~ 17:30"],
    "7-7-7-7": ["09:30 ~ 17:30", "09:30 ~ 17:30", "09:30 ~ 17:30", "09:30 ~ 17:30"],
    "8-8-4": ["09:00 ~ 18:00", "09:00 ~ 18:00", "09:00 ~ 13:00"],
    "4-8-8": ["14:00 ~ 18:00", "09:00 ~ 18:00", "09:00 ~ 18:00"],
    "8-8-8": ["09:00 ~ 18:00", "09:00 ~ 18:00", "09:00 ~ 18:00"],
    "8-8-8-8-8": ["09:00 ~ 18:00", "09:00 ~ 18:00", "09:00 ~ 18:00", "09:00 ~ 18:00", "09:00 ~ 18:00"]
    }


def get_today_cell_num(ws):
    ''' 오늘 날짜에 해당하는 셀 번호 구하기'''
    for row in ws:
        for cell in row:
            # 현재 워크시트의 전체 셀을 돌며, 현재 날짜와 cell 날짜가 같다면 해당 셀 번호 저장
            try:
                if datetime.now().date() == cell.value.date():
                    current_date_cell = cell
                    break
            except:
                pass
    return current_date_cell # ex: <Cell '8월'.N50>

def get_week_end_row(ws, start_date_cell):
    '''이번주의 마지막 행 번호 구하기'''
    week_end_row = start_date_cell.row
    # 현재 row에 날짜가 있거나, 현재 row, K열에 END_ROW_VALUE가 있다면 바로 그 위 행이 week_end_row
    while True:
        week_end_row = week_end_row + 1
        if type(ws[start_date_cell.column][week_end_row].value) == datetime or ws['K'][week_end_row].value == END_ROW_VALUE:
            week_end_row -= 1
            break
    return week_end_row

def append_upcoming_edu_rows_in_upcoming_edu_rows(ws, week_start_row, week_end_row):
    '''이번주에 시작할 과정이 있는 행을 찾아 upcoming_edu_rows 리스트에 append'''
    global day_of_week_columns_index
    for row in range(week_start_row, week_end_row):
        if ws[DAY_OF_WEEK_COLUMNS[day_of_week_columns_index]][row].value == None:
            while day_of_week_columns_index <= 3: # 월~목인 동안
                day_of_week_columns_index += 1
                if ws[DAY_OF_WEEK_COLUMNS[day_of_week_columns_index]][row].value != None:
                    upcoming_edu_rows.append((ws, row))
                    break
            day_of_week_columns_index = day_of_week_columns_index_fixed

def get_next_week_ws(current_ws, week_end_row):
    '''다음주에 해당하는 ws 구하기'''
    # 다음주에 해당하는 날짜가 있다면 next_week_ws는 현재의 ws와 동일
    if type(current_ws[MONDAY_COLUMN][week_end_row+1].value) == datetime:
        next_week_ws = current_ws
    # 그렇지 않고 다음주에 해당하는 날짜가 없다면 next_week_ws는 다음달 ws
    elif current_ws[MONDAY_COLUMN][week_end_row+1].value == END_ROW_VALUE:
         current_ws_name = current_month
         next_month = int(current_ws_name[0]) + 1
         next_week_ws_name = str(next_month) + '월'
         next_week_ws = schedule_workbook[next_week_ws_name]
    return next_week_ws

def get_next_week_row(current_ws, week_end_row):
    '''다음주 행 번호 구하기'''
    if type(current_ws[MONDAY_COLUMN][week_end_row+1].value) == datetime:
        next_week_row = week_end_row + 1
    elif current_ws[MONDAY_COLUMN][week_end_row+1].value == END_ROW_VALUE:
        next_week_row = FIRST_WEEK_ROW
    return next_week_row

def append_next_week_upcoming_edu_rows_in_upcoming_edu_rows(next_week_ws, next_week_row):
    '''다음주에 시작할 과정이 있는 행을 찾아 upcoming_edu_rows 리스트에 append'''
    next_week_start_row = next_week_row
    next_week_monday_cell = next_week_ws[MONDAY_COLUMN][next_week_row]
    next_week_end_row = get_week_end_row(next_week_ws, next_week_monday_cell)
    # 다음주 시작행부터 종료행까지 돌며 강사가 없는 경우를 제외하고 모두 upcoming_edu_rows에 추가
    for row in range(next_week_start_row, next_week_end_row):
        have_to_append = False
        # 만약 주중 강사가 없을 경우 append에서 제외시키는 조건 추가
        for column in WEEKDAY_COLUMNS:
            instructor = next_week_ws[column][row].value
            if instructor != None:
                have_to_append = True
        if have_to_append:
            upcoming_edu_rows.append((next_week_ws, row))

def get_selected_edu_sections_row():
    ''''upcoming_edu_rows에서 input parameter로 받아온 과정 부문(EDU_SECTION)에 해당하는 행만 구하기'''
    selected_edu_sections_row = []
    for ws, row in upcoming_edu_rows:
        if ws[EDU_SECTION_COLUMN][row].value == EDU_SECTION:
            selected_edu_sections_row.append((ws, row))
    return selected_edu_sections_row

# for work sheet
def append_ws_in_df_scripts():
    '''selected_edu_secitons에서 work sheet를 추출하여 df_scripts에 저장'''
    temp_list = []
    for ws, row in selected_edu_sections_row:
        temp_list.append(ws)
    df_scripts['work_sheet'] = temp_list

# for rows
def append_edu_row_in_df_scripts():
    '''selected_edu_secitons에서 row를 추출하여 df_scripts에 저장'''
    temp_list = []
    for ws, row in selected_edu_sections_row:
        temp_list.append(row)
    df_scripts['edu_row'] = temp_list

# for education name
def append_edu_name_in_df_scripts():
    '''selected_edu_secitons로부터 과정명을 구하여 df_scripts에 저장'''
    temp_list = []
    for ws, row in selected_edu_sections_row:
        temp_list.append(ws[EDU_COLUMN][row].value)
    df_scripts['edu_name'] = temp_list

# for education room
def append_edu_room_in_df_scripts():
    '''selected_edu_secitons로부터 강의실을 구하여 df_scripts에 저장'''
    temp_list = []
    for ws, row in selected_edu_sections_row:
        if ws[CHANGED_EDU_ROOM_COLUMN][row].value == None:
            # 강의장이 숫자면 서울 강의장, 문자면 지역 본부
            if type(ws[EDU_ROOM_COLUMN][row].value) == int:
                temp_list.append('KPC 서울 본부 '+ str(ws[EDU_ROOM_COLUMN][row].value) + '호 강의장')
            elif type(ws[EDU_ROOM_COLUMN][row].value) == str:
                temp_list.append('KPC ' + ws[EDU_ROOM_COLUMN][row].value + ' 본부')
        # 강의장이 변경되었다면 변경된 강의장을 저장
        elif ws[CHANGED_EDU_ROOM_COLUMN][row].value != None:
            if type(ws[CHANGED_EDU_ROOM_COLUMN][row].value) == int:
                temp_list.append('KPC 서울 본부 '+ str(ws[CHANGED_EDU_ROOM_COLUMN][row].value) + '호 강의장')
            elif type(ws[CHANGED_EDU_ROOM_COLUMN][row].value) == str:
                temp_list.append('KPC ' + ws[CHANGED_EDU_ROOM_COLUMN][row].value + ' 본부')
    df_scripts['edu_room'] = temp_list

def get_week_start_row(ws, row_num):
    '''현재 행이 포함되어 있는 week의 날짜가 있는 행번호 구하기'''
    while type(ws[MONDAY_COLUMN][row_num].value) != datetime:
        row_num -= 1
    return row_num
    
def append_start_end_date_in_df_scripts():
    '''일정 도출을 위해 과정 시작날짜, 종료날짜 구하여 df_scripts에 저장'''
    temp_start_list = []
    temp_end_list = []
    for index, row in df_scripts.iterrows():
        find_edu_start_date = False
        find_edu_end_date = False
        row_num = row['edu_row']
        ws = row['work_sheet']
        week_start_row = get_week_start_row(ws, row_num)
        for column in WEEKDAY_COLUMNS:
            # 교육 시작일 찾기
            if ws[column][row_num].value != None and find_edu_start_date == False:
                edu_start_date = ws[column][week_start_row].value
                find_edu_start_date = True
                temp_start_list.append(edu_start_date)
            # 교육 종료일 찾기
            elif ws[column][row_num].value == None and find_edu_end_date == False and find_edu_start_date == True:
                # 빈 셀의 바로 왼쪽이 교육 종료일
                edu_end_column = WEEKDAY_COLUMNS[WEEKDAY_COLUMNS.index(column)-1]
                edu_end_date = ws[edu_end_column][week_start_row].value
                find_edu_end_date = True
                temp_end_list.append(edu_end_date)
            # 종료일이 금요일인 경우
            elif ws[column][row_num].value != None and find_edu_end_date == False and find_edu_start_date == True and column == 'O':
                edu_end_date = ws[column][week_start_row].value
                find_edu_end_date = True
                temp_end_list.append(edu_end_date)
    df_scripts['edu_start_date'] = temp_start_list
    df_scripts['edu_end_date'] = temp_end_list

# for instructors
def append_instructors_in_df_scripts():
    '''강사를 df_scripts에 저장'''
    temp_list = []
    for index, row in df_scripts.iterrows():
        row_num = row['edu_row']
        ws = row['work_sheet']
        instructors_in_one_edu = []
        for column in WEEKDAY_COLUMNS:
            if ws[column][row_num].value != None:
                instructor = ws[column][row_num].value
                # '/' 삭제
                instructor = instructor.replace("/", "")
                instructors_in_one_edu.append(instructor)
        temp_list.append(instructors_in_one_edu)
    df_scripts['instructors'] = temp_list

# for full education date
def append_full_edu_date_in_df_scripts():
    '''과정 시작일부터 종료일까지 MM.DD(WEEKDAY) 형태로 df_scripts에 저장'''
    temp_list = []
    weekday_dict = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
    for index, row in df_scripts.iterrows():
        full_dates = []
        date_difference = (row['edu_end_date']-row['edu_start_date']).days + 1
        edu_date = row['edu_start_date']
        # 시작일부터 종료일까지 full_date 형식으로 저장
        for i in range(date_difference):
            month = '{:02d}'.format(edu_date.month)
            day = '{:02d}'.format(edu_date.day)
            weekday = weekday_dict[edu_date.weekday()]
            full_date = str(month) + '.' +str(day) + '(' + weekday +')'
            full_dates.append(full_date)
            # 다음날
            edu_date = edu_date + timedelta(days=1)
        temp_list.append(full_dates)
    df_scripts['full_dates'] = temp_list

# for education time
def append_edu_time_in_df_scripts():
    '''7-7-4 형태의 시간 양식으로부터 교육 시간을 구하여 df_scripts에 저장
        1. 첫 숫자를 가져온다.
        2. 그 다음 문자가 '-'이면 다음 숫자까지 가져온다.
        3. 그 다음 문자가 '-'가 아닐 때까지 반복한다.
    '''
    edu_time = []
    temp_list = []
    for index, row in df_scripts.iterrows():
        row_num = row['edu_row']
        ws = row['work_sheet']
        time_strings = ws[TIME_COLUMN][row_num].value
        string_index = 0
        for index in range(len(time_strings)):
            # time으로만 이루어진 string의 마지막 index일 때
            if index+1 == len(time_strings) and time_strings[index-1] == '-':
                string_index = index
                break
            # time 이외의 문구가 있는 string의 마지막 index일 때
            elif index+1 == len(time_strings) and time_strings[index-1] != '-':
                continue
            elif time_strings[index+1] == '-':
                continue
            elif time_strings[index] == '-':
                continue
            elif time_strings[index+1] != '-':
                string_index = index
                break
        edu_time.append(time_strings[:string_index+1])
    # DETAILED_EDU_TIME dict로부터 시간 받아오기
    for time in edu_time:
        detailed_edu_time = DETAILED_EDU_TIME[time]
        temp_list.append(detailed_edu_time)
    df_scripts['edu_time'] = temp_list  

def has_only_one_instructor(instructor_list):
    '''instructor_list에 강사가 한명인지 여부 판단'''
    instructor_set = set(instructor_list)
    if len(instructor_set) == 1:
        return True
    else:
        return False

# for detailed date time and instructor
def append_detailed_date_time_instructor_in_df_scripts():
    '''세부 강의 시간 및 배정 강사 스크립트를 df_scripts에 저장'''
    temp_list = []
    for index, row in df_scripts.iterrows():
        detailed_date_time_instructor = []
        full_date = row['full_dates']
        edu_time = row['edu_time']
        instructors = row['instructors']
        for index in range(len(full_date)):
            # 강사가 한명일 경우 강사 이름은 제외
            if has_only_one_instructor(instructors):
                detailed_date_time_instructor_scripts = full_date[index] + " " + edu_time[index]
            # 강사가 여러명일 경우 해당 일정 배정 강사 기재
            else:
                detailed_date_time_instructor_scripts = full_date[index] + " " + edu_time[index] + " : " + instructors[index] +" 지도위원님"
            detailed_date_time_instructor.append(detailed_date_time_instructor_scripts)
        temp_list.append(detailed_date_time_instructor)
    df_scripts['detailed_date_time_instructor'] = temp_list

# for full mail scripts
def append_full_mail_script_in_df_scripts():
    '''실제 메일 스크립트를 df_scripts에 저장'''
    temp_list = []
    for index, row in df_scripts.iterrows():
        temp_string = GREETING_START +'\n\n' # 인사말
        temp_string += '1. 과정명 : ' + row['edu_name'] + '\n'
        temp_string += '2. 일정 : ' + row['full_dates'][0] + ' ~ ' + row['full_dates'][-1] + '\n'
        temp_string += '3. 강의장소 : ' + row['edu_room'] + '\n'
        temp_string += '4. 강의시간: \n'
        for detailed_date_time in row['detailed_date_time_instructor']:
            temp_string += '  ' + detailed_date_time + '\n'
        temp_string += '5. 준비사항 : 교안 파일 및 참고 자료는 USB에 담아서 준비헤주시기 바랍니다.\n'
        temp_string += '6. 수강생명단 : 첨부파일 참조\n\n'
        temp_string += GREETING_END # 끝인사
        temp_list.append(temp_string)
    df_scripts['full_mail_scripts'] = temp_list

def append_info_in_df_scripts(df_scripts):
    append_ws_in_df_scripts()
    append_edu_row_in_df_scripts()
    append_edu_name_in_df_scripts()
    append_edu_room_in_df_scripts()
    append_start_end_date_in_df_scripts()
    append_instructors_in_df_scripts()
    append_full_edu_date_in_df_scripts()   
    append_edu_time_in_df_scripts()
    append_detailed_date_time_instructor_in_df_scripts()
    append_full_mail_script_in_df_scripts()

def get_current_month_worksheet(schedule_workbook):
    '''현재 월에 해당하는 워크시트 불러오기'''
    current_month = str(datetime.now().month) + '월'
    ws = schedule_workbook[current_month]
    return ws

def adjust_mail_sciprts_cell_attributes():
    mail_scripts_workbook = openpyxl.load_workbook('mail_script.xlsx')
    mail_scripts_worksheet = mail_scripts_workbook['Sheet1']
    # 열 너비 조정
    mail_scripts_worksheet.column_dimensions['B'].width = 70
    # 줄바꿈, 폰트 및 사이즈 조정
    row_count = 1
    cell_num = 'B' + str(row_count)
    while mail_scripts_worksheet[cell_num].value != None:
        mail_scripts_worksheet[cell_num].alignment = Alignment(horizontal='justify', wrap_text=True)
        mail_scripts_worksheet[cell_num].font = Font(name='맑은 고딕', size=10)
        row_count += 1
        cell_num = 'B' + str(row_count)
    # 행 높이 조정
    for i in range(mail_scripts_worksheet.max_row):
        mail_scripts_worksheet.row_dimensions[i+1].height = 70
    # 저장
    mail_scripts_workbook.save("mail_script.xlsx")
    
if __name__ == "__main__":
    schedule_workbook = openpyxl.load_workbook(SCHEDULE_FILE)
    ws = get_current_month_worksheet(schedule_workbook)

    upcoming_edu_rows = []

    # 이번주 과정의 행 시작과 끝 구하기
    current_date_cell = get_today_cell_num(ws)
    week_start_row = current_date_cell.row
    week_end_row = get_week_end_row(ws, current_date_cell)
    
    # 다음주 과정의 워크시트와 행 구하기
    next_week_ws = get_next_week_ws(ws, week_end_row)
    next_week_row = get_next_week_row(ws, week_end_row)

    day_of_week_columns_index = DAY_OF_WEEK_COLUMNS.index(current_date_cell.column)
    day_of_week_columns_index_fixed = day_of_week_columns_index

    append_upcoming_edu_rows_in_upcoming_edu_rows(ws, week_start_row, week_end_row) # 이번주 개강할 과정
    append_next_week_upcoming_edu_rows_in_upcoming_edu_rows(next_week_ws, next_week_row) # 다음주 개강할 과정
    
    # 해당 부문 선별
    selected_edu_sections_row = get_selected_edu_sections_row()

    df_scripts = pd.DataFrame()
    append_info_in_df_scripts(df_scripts)

    # output 파일로 저장
    df_scripts['full_mail_scripts'].to_excel(OUTPUT_SCRIPT_FILE)
    
    adjust_mail_sciprts_cell_attributes()

    # 남은 일: outputfile에 과정 추가